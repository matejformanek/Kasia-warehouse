"""Stock-mutation + audit service for Movement create / edit.

The functions below are the single write path for Movement, MovementLine,
Stock, and MovementAudit. Admin (and future views) call them inside an
atomic block; signals are deliberately avoided so the call graph stays
visible.

See plan: context/state.md § Next item 2 (Pass 1 of 2).
Schema invariants: decisions 0021 + 0035 (audit shape), 0030 (kind enum),
0028 (mass-only), 0001 (šarže optional), 0003 (NUMERIC(10,3)).

Note: select_for_update() is a silent no-op on SQLite. Real concurrency
safety arrives once every code path runs against Postgres. Acceptable
for MVP at ~6 users (per .claude/rules/right-sized-for-small-business.md).
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import Any

from django.core.exceptions import ValidationError
from django.core.mail import EmailMessage
from django.db import IntegrityError, transaction
from django.template.loader import render_to_string

from .models import (
    Branch,
    Customer,
    DodaciList,
    DodaciListEmailLog,
    DodaciListNumberSequence,
    MixingJob,
    MixingJobLine,
    Movement,
    MovementAudit,
    MovementLine,
    PlannedTransfer,
    Product,
    RecipeComponent,
    Settings,
    Stock,
    StockThresholdOverride,
    Supplier,
)

_MOVEMENT_AUDITABLE_FIELDS = ("kind", "branch", "date_issued", "odberatel", "dodavatel", "note")
_LINE_AUDITABLE_FIELDS = ("product", "quantity_kg", "sarze", "expiry", "note")


def _render(value: Any) -> str:
    """String-render a field value for the audit log. Empty for None."""
    if value is None:
        return ""
    return str(value)


def _line_summary(line: MovementLine) -> str:
    """Human-readable one-line snapshot of a MovementLine for line_added /
    line_removed audit entries."""
    parts = [f"{line.product} {line.quantity_kg} kg"]
    if line.sarze:
        parts.append(f"šarže {line.sarze}")
    if line.expiry:
        parts.append(f"exp {line.expiry.isoformat()}")
    if line.note:
        parts.append(line.note)
    return " · ".join(parts)


def _apply_line_to_stock(line: MovementLine, *, direction: int) -> None:
    """Mutate the (product, branch) Stock row by `direction * line.quantity_kg`.
    Raises ValidationError if the resulting quantity would go negative."""
    stock, _ = Stock.objects.select_for_update().get_or_create(
        product=line.product,
        branch_id=line.movement.branch_id,
        defaults={"quantity": Decimal("0.000")},
    )
    stock.quantity = stock.quantity + Decimal(direction) * line.quantity_kg
    try:
        # Nested savepoint so a failed CHECK CONSTRAINT rolls back to here,
        # not the whole outer transaction — letting us convert to a friendly
        # ValidationError without leaving the outer atomic block broken.
        with transaction.atomic():
            stock.save()
    except IntegrityError as exc:
        raise ValidationError(
            {
                "quantity_kg": (
                    f"Skladová zásoba by klesla pod nulu "
                    f"(produkt: {line.product}, pobočka: {line.movement.branch.code})."
                )
            }
        ) from exc


def apply_movement(
    *,
    movement: Movement,
    lines: list[MovementLine],
    user,
) -> Movement:
    """Create a Movement + its lines atomically and mutate Stock.

    No MovementAudit rows are written — the Movement row itself (with
    created_at / created_by) is the creation record.
    """
    if not lines:
        raise ValidationError({"lines": "Pohyb musí mít alespoň jednu položku."})

    direction = 1 if movement.kind == Movement.Kind.PRIJEM else -1

    # Internal-counterparty pohyby (e.g. míchání-job consume per
    # decision 0039) bypass the dodák PDF + e-mail path entirely.
    # apply_movement still writes the Movement + decrements stock; the
    # operator-facing surface treats these rows like any other movement
    # except they aren't paired with a DodaciList.
    is_internal_vydej = (
        movement.kind == Movement.Kind.VYDEJ
        and movement.odberatel is not None
        and movement.odberatel.is_internal
    )

    if movement.kind == Movement.Kind.VYDEJ and not is_internal_vydej:
        _assert_recipients_set()

    with transaction.atomic():
        movement.created_by = user
        movement.full_clean()
        movement.save()
        for line in lines:
            line.movement = movement
            line.full_clean()
            line.save()
            _apply_line_to_stock(line, direction=direction)

        if movement.kind == Movement.Kind.VYDEJ and not is_internal_vydej:
            dodaci_list = _create_dodaci_list_for_movement(movement)
            pdf_bytes = render_dodaci_list_pdf(dodaci_list)
            transaction.on_commit(
                lambda dl=dodaci_list, pdf=pdf_bytes: send_dodaci_list_email(
                    dodaci_list=dl,
                    trigger_reason="vystavení",
                    pdf_bytes=pdf,
                )
            )

        return movement


def edit_movement(
    *,
    movement: Movement,
    changes: dict[str, Any],
    line_changes: list[dict[str, Any]],
    reason: str,
    user,
) -> Movement:
    """Edit an existing Movement atomically.

    `changes` is a flat dict of Movement-level field → new value (only
    fields that actually change). `line_changes` is a list of per-line
    operations:

        {"op": "update", "line_id": int, "fields": {field: new_value, ...}}
        {"op": "add", "fields": {"product": prod, "quantity_kg": Dec, ...}}
        {"op": "remove", "line_id": int}

    Writes one MovementAudit row per *changed* movement field, per
    *changed* line field, and per add / remove event. Recomputes Stock
    deltas; rolls back the whole edit if stock would go negative.
    """
    if not reason or not reason.strip():
        raise ValidationError({"reason": "Důvod úpravy je povinný."})
    if "kind" in changes:
        raise ValidationError({"kind": "Druh pohybu nelze změnit úpravou; vytvořte nový pohyb."})

    direction = 1 if movement.kind == Movement.Kind.PRIJEM else -1

    with transaction.atomic():
        audit_rows: list[MovementAudit] = []

        # ---- Movement-level field changes ------------------------------------
        for field, new_value in changes.items():
            if field not in _MOVEMENT_AUDITABLE_FIELDS:
                raise ValidationError({field: f"Pole '{field}' nelze upravit."})
            old_value = getattr(movement, field)
            if old_value == new_value:
                continue
            audit_rows.append(
                MovementAudit(
                    movement=movement,
                    edited_by=user,
                    reason=reason,
                    target_kind=MovementAudit.TargetKind.MOVEMENT,
                    line_id=None,
                    event=MovementAudit.Event.FIELD_CHANGED,
                    field=field,
                    old_value=_render(old_value),
                    new_value=_render(new_value),
                )
            )
            setattr(movement, field, new_value)

        if any(c for c in changes if c in _MOVEMENT_AUDITABLE_FIELDS):
            movement.full_clean()
            movement.save()

        # ---- Per-line changes ------------------------------------------------
        for op in line_changes:
            action = op.get("op")
            if action == "update":
                line = MovementLine.objects.select_related("movement", "product").get(
                    pk=op["line_id"], movement=movement
                )
                fields = op.get("fields", {})
                # Snapshot the stock-relevant pair before mutating.
                old_product = line.product
                old_quantity = line.quantity_kg

                changed_any = False
                stock_relevant_change = False
                for field, new_value in fields.items():
                    if field not in _LINE_AUDITABLE_FIELDS:
                        raise ValidationError({field: f"Pole položky '{field}' nelze upravit."})
                    old_value = getattr(line, field)
                    if old_value == new_value:
                        continue
                    changed_any = True
                    if field in ("quantity_kg", "product"):
                        stock_relevant_change = True
                    audit_rows.append(
                        MovementAudit(
                            movement=movement,
                            edited_by=user,
                            reason=reason,
                            target_kind=MovementAudit.TargetKind.LINE,
                            line_id=line.pk,
                            event=MovementAudit.Event.FIELD_CHANGED,
                            field=field,
                            old_value=_render(old_value),
                            new_value=_render(new_value),
                        )
                    )
                    setattr(line, field, new_value)

                if not changed_any:
                    continue

                if stock_relevant_change:
                    reverse_line = MovementLine(
                        movement=movement,
                        product=old_product,
                        quantity_kg=old_quantity,
                    )
                    _apply_line_to_stock(reverse_line, direction=-direction)
                    line.full_clean()
                    line.save()
                    _apply_line_to_stock(line, direction=direction)
                else:
                    line.full_clean()
                    line.save()

            elif action == "add":
                fields = op.get("fields", {})
                line = MovementLine(movement=movement, **fields)
                line.full_clean()
                line.save()
                _apply_line_to_stock(line, direction=direction)
                audit_rows.append(
                    MovementAudit(
                        movement=movement,
                        edited_by=user,
                        reason=reason,
                        target_kind=MovementAudit.TargetKind.LINE,
                        line_id=line.pk,
                        event=MovementAudit.Event.LINE_ADDED,
                        field="",
                        old_value="",
                        new_value=_line_summary(line),
                    )
                )

            elif action == "remove":
                line = MovementLine.objects.select_related("movement", "product").get(
                    pk=op["line_id"], movement=movement
                )
                summary = _line_summary(line)
                _apply_line_to_stock(line, direction=-direction)
                line_pk = line.pk
                line.delete()
                audit_rows.append(
                    MovementAudit(
                        movement=movement,
                        edited_by=user,
                        reason=reason,
                        target_kind=MovementAudit.TargetKind.LINE,
                        line_id=line_pk,
                        event=MovementAudit.Event.LINE_REMOVED,
                        field="",
                        old_value=summary,
                        new_value="",
                    )
                )

            else:
                raise ValidationError({"op": f"Neznámá operace položky: {action!r}."})

        MovementAudit.objects.bulk_create(audit_rows)

        # Per decision 0007: a movement edit on a posted dodák bumps the
        # internal version counter, re-renders the PDF against current
        # data + template, and re-sends with an [OPRAVA] subject. The
        # send itself runs on commit so a rollback of the outer atomic
        # block (e.g. a stock overdraw later) skips the e-mail entirely.
        dodaci_list = DodaciList.objects.filter(movement=movement).first()
        if dodaci_list is not None:
            _assert_recipients_set()
            dodaci_list.current_version += 1
            dodaci_list.save(update_fields=["current_version"])
            pdf_bytes = render_dodaci_list_pdf(dodaci_list)
            transaction.on_commit(
                lambda dl=dodaci_list, pdf=pdf_bytes, r=reason: send_dodaci_list_email(
                    dodaci_list=dl,
                    trigger_reason=f"oprava: {r}",
                    pdf_bytes=pdf,
                )
            )

        return movement


# ---------------------------------------------------------------------------
# Dodací list services (per 0007 / 0008 / 0017 / 0019 / 0031 / 0036 / 0037)
# ---------------------------------------------------------------------------


def _assert_recipients_set() -> None:
    """Refuse to start a vydej apply / edit if Settings recipients are blank.

    Per 0031 every dodák goes to the fixed (Petr, Karolína) pair from
    Settings; the seed migration leaves them empty intentionally so an
    operator fills them on first run. We check before doing any DB work
    that the on-commit hook can't roll back.
    """
    s = Settings.load()
    if not s.recipient_petr or not s.recipient_karolina:
        raise ValidationError(
            {
                "recipients": (
                    "V nastavení chybí příjemci dodacího listu "
                    "(Petr a Karolína). Doplňte je v Nastavení před výdejem."
                )
            }
        )


def _reserve_dodak_number(*, branch, year: int) -> int:
    """Allocate the next per-(branch, year) counter under SELECT … FOR UPDATE.

    Must be called inside the caller's transaction.atomic() block — the
    row lock is released at commit.
    """
    seq, _ = (
        DodaciListNumberSequence.objects.select_for_update().get_or_create(
            branch=branch, year=year, defaults={"last_counter": 0}
        )
    )
    seq.last_counter += 1
    seq.save(update_fields=["last_counter"])
    return seq.last_counter


def _create_dodaci_list_for_movement(movement: Movement) -> DodaciList:
    """Insert one DodaciList row for a vydej Movement.

    Inside the caller's atomic block. Computes cislo from
    (branch.code, year, counter) per 0008.
    """
    year = movement.date_issued.year
    counter = _reserve_dodak_number(branch=movement.branch, year=year)
    cislo = f"{movement.branch.code}-{year}-{counter:04d}"
    return DodaciList.objects.create(
        movement=movement,
        branch=movement.branch,
        odberatel=movement.odberatel,
        date_issued=movement.date_issued,
        year_issued=year,
        counter=counter,
        cislo=cislo,
        current_version=1,
        created_by=movement.created_by,
    )


def render_dodaci_list_pdf(dodaci_list: DodaciList) -> bytes:
    """Render the dodák PDF via WeasyPrint per 0017.

    Template is templates/inventory/dodaci_list.html with embedded CSS
    Paged Media. Always re-renders against current Settings + current
    Customer (per 0007 / 0036).
    """
    # Imported lazily — keeps module import cheap for non-PDF callers
    # (admin pages, tests touching services without rendering).
    from pathlib import Path

    from django.conf import settings as django_settings
    from weasyprint import HTML

    lines = list(dodaci_list.movement.lines.all().order_by("id"))

    # WeasyPrint resolves images from disk. When Settings.logo is empty,
    # fall back to the bundled brand mark at
    # kasia/static/brand/kasia-logo.jpg. The template gets an absolute
    # file:// URL it can use directly.
    bundled_logo = Path(django_settings.BASE_DIR) / "kasia" / "static" / "brand" / "kasia-logo.jpg"
    default_logo_url = f"file://{bundled_logo}" if bundled_logo.exists() else ""

    html_string = render_to_string(
        "inventory/dodaci_list.html",
        {
            "dodaci_list": dodaci_list,
            "movement": dodaci_list.movement,
            "lines": lines,
            "show_sarze": any(line.sarze for line in lines),
            "show_note": any(line.note for line in lines),
            "settings": Settings.load(),
            "default_logo_url": default_logo_url,
        },
    )
    return HTML(string=html_string).write_pdf()


def _substitute_template(text: str, dodaci_list: DodaciList) -> str:
    """Substitute the screen-14 placeholders. Reason placeholder is only
    used in the oprava body; the caller is expected to append the
    operator reason via trigger_reason for the e-mail log."""
    return (
        text
        .replace("<číslo>", dodaci_list.cislo)
        .replace("<datum>", dodaci_list.date_issued.strftime("%d. %m. %Y"))
    )


def send_dodaci_list_email(
    *,
    dodaci_list: DodaciList,
    trigger_reason: str,
    pdf_bytes: bytes,
) -> DodaciListEmailLog:
    """Send one dodák e-mail to the fixed (Petr, Karolína) pair per 0031.

    Wrapped in try/except per 0019: a send failure writes a FAILED log
    row and returns it; it does NOT re-raise. The výdej / oprava write
    that triggered the send is already committed.
    """
    s = Settings.load()
    recipients = [s.recipient_petr, s.recipient_karolina]
    recipients_joined = ", ".join(recipients)
    is_oprava = trigger_reason.startswith("oprava")
    subject_template = (
        s.template_oprava_subject if is_oprava else s.template_initial_subject
    )
    body_template = s.template_oprava_body if is_oprava else s.template_initial_body
    subject = _substitute_template(subject_template, dodaci_list)
    body = _substitute_template(body_template, dodaci_list)
    if is_oprava:
        operator_reason = trigger_reason[len("oprava:") :].strip()
        body = body.replace("<text zdůvodnění od operátorky>", operator_reason)

    from_email = (
        f"{s.email_from_name} <{s.email_from_address}>"
        if s.email_from_name and s.email_from_address
        else (s.email_from_address or None)
    )

    msg = EmailMessage(
        subject=subject,
        body=body,
        from_email=from_email,
        to=recipients,
    )
    msg.attach(f"{dodaci_list.cislo}.pdf", pdf_bytes, "application/pdf")

    try:
        msg.send(fail_silently=False)
    except Exception as exc:
        return DodaciListEmailLog.objects.create(
            dodaci_list=dodaci_list,
            version=dodaci_list.current_version,
            recipients=recipients_joined,
            trigger_reason=trigger_reason,
            status=DodaciListEmailLog.Status.FAILED,
            error_message=str(exc),
        )

    return DodaciListEmailLog.objects.create(
        dodaci_list=dodaci_list,
        version=dodaci_list.current_version,
        recipients=recipients_joined,
        trigger_reason=trigger_reason,
        status=DodaciListEmailLog.Status.SENT,
    )


# ---------------------------------------------------------------------------
# Mixing job services (screen 15, per decision 0039)
# ---------------------------------------------------------------------------


def _micharna_customer() -> Customer:
    return Customer.objects.get(name="Míchárna", is_internal=True)


def _micharna_supplier() -> Supplier:
    return Supplier.objects.get(name="Míchárna", is_internal=True)


def plan_mixing_job(
    *,
    branch,
    mixture: Product,
    target_qty: Decimal,
    user,
    planned_for=None,
    note: str = "",
) -> MixingJob:
    """Create a PLANNED MixingJob without touching Stock.

    Per [0044](../context/decisions/0044-reservations-planned-states.md):
    PLANNED jobs reserve stock via `reserved_kg()` (their
    `MixingJobLine.derived_qty` rows feed the reservation total) but
    do not decrement `Stock.quantity`. The transition to RUNNING — via
    `start_mixing_job(job=<planned>)` — is the moment stock is
    actually consumed.

    Snapshots the recipe at the moment of planning so a future recipe
    edit doesn't retroactively change a planned job's reservation.
    """
    if mixture.kind != Product.Kind.MIXTURE:
        raise ValidationError({"mixture": "Vybraný produkt není směs."})
    if target_qty is None or target_qty <= 0:
        raise ValidationError(
            {"target_qty": "Cílové množství musí být větší než 0."}
        )

    recipe = list(
        RecipeComponent.objects.filter(mixture_product=mixture)
        .select_related("component_product")
        .order_by("component_product__name_cs")
    )
    if not recipe:
        raise ValidationError({"mixture": "Směs nemá vyplněnou recepturu."})

    with transaction.atomic():
        job = MixingJob.objects.create(
            branch=branch,
            mixture=mixture,
            target_qty=target_qty,
            state=MixingJob.State.PLANNED,
            planned_for=planned_for,
            created_by=user,
            note=note,
        )
        for rc in recipe:
            derived = (target_qty * rc.ratio).quantize(Decimal("0.001"))
            if derived <= 0:
                raise ValidationError(
                    {
                        "target_qty": (
                            f"Odvozené množství pro {rc.component_product} "
                            f"je 0; zvolte větší cíl."
                        )
                    }
                )
            MixingJobLine.objects.create(
                mixing_job=job,
                component_product=rc.component_product,
                ratio_at_start=rc.ratio,
                derived_qty=derived,
                actual_qty=derived,
            )
        return job


def start_mixing_job(
    *,
    branch=None,
    mixture: Product | None = None,
    target_qty: Decimal | None = None,
    user,
    as_of=None,
    note: str = "",
    sarze_by_component: dict | None = None,
    job: MixingJob | None = None,
) -> MixingJob:
    """Snapshot the recipe at the current ratios, write the consume
    Movement (kind=vydej, odberatel=Míchárna internal) atomically, and
    return a running MixingJob.

    Two entry shapes, per
    [0044](../context/decisions/0044-reservations-planned-states.md):

    - **Fresh start (one-shot):** caller passes branch, mixture,
      target_qty. We create the RUNNING MixingJob + consume Movement
      in one atomic block.
    - **From a PLANNED job:** caller passes `job=<planned MixingJob>`.
      We assert `job.state == PLANNED`, snapshot already exists on
      `MixingJobLine` rows, transition state PLANNED → RUNNING, write
      the consume Movement, link `consume_movement` to the job.

    Raises ValidationError on:
    - mixture without recipe (fresh-start);
    - target_qty <= 0 (fresh-start);
    - any component's stock would go negative at this branch.

    Per 0039: ratios snapshotted at start; future recipe edits don't
    touch in-flight jobs. Stock-overdraw refusal hits via the existing
    `_apply_line_to_stock` invariant.
    """
    from datetime import date as _date

    sarze_by_component = sarze_by_component or {}
    date_issued = as_of.date() if hasattr(as_of, "date") else (
        as_of if as_of is not None else _date.today()
    )

    if job is not None:
        # PLANNED → RUNNING path.
        if job.state != MixingJob.State.PLANNED:
            raise ValidationError(
                {"state": "Spustit lze pouze plánovanou dávku."}
            )
        branch = job.branch
        mixture = job.mixture
        target_qty = job.target_qty
        existing_lines = list(
            job.lines.select_related("component_product").order_by(
                "component_product__name_cs"
            )
        )
        if not existing_lines:
            raise ValidationError(
                {"mixture": "Plánovaná dávka nemá žádné položky."}
            )

        with transaction.atomic():
            consume_movement = Movement(
                branch=branch,
                kind=Movement.Kind.VYDEJ,
                date_issued=date_issued,
                odberatel=_micharna_customer(),
                note=(
                    f"Míchání směsi {mixture.name_cs} ({target_qty} kg). "
                    f"{job.note or note}".strip()
                ),
            )
            consume_lines = [
                MovementLine(
                    product=jl.component_product,
                    quantity_kg=jl.derived_qty,
                    sarze=sarze_by_component.get(jl.component_product_id, jl.sarze or ""),
                )
                for jl in existing_lines
            ]
            apply_movement(
                movement=consume_movement, lines=consume_lines, user=user
            )
            # Mirror sarze input back onto the MixingJobLine rows.
            for jl in existing_lines:
                new_sarze = sarze_by_component.get(jl.component_product_id)
                if new_sarze and new_sarze != jl.sarze:
                    jl.sarze = new_sarze
                    jl.save(update_fields=["sarze"])
            job.state = MixingJob.State.RUNNING
            job.consume_movement = consume_movement
            job.save(update_fields=["state", "consume_movement"])
            return job

    # Fresh-start path.
    if mixture is None or branch is None:
        raise ValidationError(
            {"mixture": "Pobočka a směs jsou povinné."}
        )
    if mixture.kind != Product.Kind.MIXTURE:
        raise ValidationError(
            {"mixture": "Vybraný produkt není směs."}
        )
    if target_qty is None or target_qty <= 0:
        raise ValidationError(
            {"target_qty": "Cílové množství musí být větší než 0."}
        )

    recipe = list(
        RecipeComponent.objects.filter(mixture_product=mixture)
        .select_related("component_product")
        .order_by("component_product__name_cs")
    )
    if not recipe:
        raise ValidationError(
            {"mixture": "Směs nemá vyplněnou recepturu."}
        )

    with transaction.atomic():
        # Build the consume Movement (one vydej with N lines).
        consume_movement = Movement(
            branch=branch,
            kind=Movement.Kind.VYDEJ,
            date_issued=date_issued,
            odberatel=_micharna_customer(),
            note=(
                f"Míchání směsi {mixture.name_cs} ({target_qty} kg). "
                f"{note}".strip()
            ),
        )

        consume_lines: list[MovementLine] = []
        snapshots: list[tuple[Product, Decimal, Decimal]] = []
        for rc in recipe:
            derived = (target_qty * rc.ratio).quantize(Decimal("0.001"))
            if derived <= 0:
                # Pathological: a positive ratio rounded to 0 at 3 dp.
                raise ValidationError(
                    {
                        "target_qty": (
                            f"Odvozené množství pro {rc.component_product} "
                            f"je 0; zvolte větší cíl."
                        )
                    }
                )
            consume_lines.append(
                MovementLine(
                    product=rc.component_product,
                    quantity_kg=derived,
                    sarze=sarze_by_component.get(rc.component_product_id, ""),
                )
            )
            snapshots.append((rc.component_product, rc.ratio, derived))

        apply_movement(
            movement=consume_movement, lines=consume_lines, user=user
        )

        new_job = MixingJob.objects.create(
            branch=branch,
            mixture=mixture,
            target_qty=target_qty,
            state=MixingJob.State.RUNNING,
            created_by=user,
            note=note,
            consume_movement=consume_movement,
        )
        MixingJobLine.objects.bulk_create(
            [
                MixingJobLine(
                    mixing_job=new_job,
                    component_product=component,
                    ratio_at_start=ratio,
                    derived_qty=derived,
                    actual_qty=derived,
                    sarze=sarze_by_component.get(component.pk, ""),
                )
                for component, ratio, derived in snapshots
            ]
        )
        return new_job


def finish_mixing_job(
    *,
    mixing_job: MixingJob,
    actual_produced_qty: Decimal,
    line_actuals: dict[int, Decimal] | None = None,
    user,
    as_of=None,
) -> MixingJob:
    """Write the produce Movement, persist any operator-edited actual
    consumption per component, and mark the job done.

    `line_actuals` is `{mixing_job_line_id: actual_qty}`. Missing
    entries keep the line's existing `actual_qty` (defaulted to
    `derived_qty` at start). If an actual differs from the derived,
    the consume Movement is corrected via `edit_movement` with a
    canned reason so the audit trail captures it.
    """
    from datetime import date as _date

    if mixing_job.state != MixingJob.State.RUNNING:
        raise ValidationError(
            {"state": "Lze ukončit pouze probíhající dávku."}
        )
    if actual_produced_qty is None or actual_produced_qty < 0:
        raise ValidationError(
            {
                "actual_produced_qty": (
                    "Skutečné vyrobené množství nemůže být záporné."
                )
            }
        )

    line_actuals = line_actuals or {}
    date_issued = as_of.date() if hasattr(as_of, "date") else (
        as_of if as_of is not None else _date.today()
    )

    with transaction.atomic():
        # Per-line actual edits — applied via edit_movement so the
        # stock delta + audit trail are computed by the existing service.
        line_changes = []
        consume_movement = mixing_job.consume_movement
        consume_lines_by_product = {
            ml.product_id: ml for ml in consume_movement.lines.all()
        }

        for jl_id, new_actual in line_actuals.items():
            jl = MixingJobLine.objects.get(pk=jl_id, mixing_job=mixing_job)
            new_actual = Decimal(new_actual).quantize(Decimal("0.001"))
            if new_actual <= 0:
                raise ValidationError(
                    {
                        "line_actuals": (
                            f"Spotřeba {jl.component_product} musí být > 0."
                        )
                    }
                )
            jl.actual_qty = new_actual
            jl.save(update_fields=["actual_qty"])

            consume_line = consume_lines_by_product.get(jl.component_product_id)
            if consume_line is not None and consume_line.quantity_kg != new_actual:
                line_changes.append(
                    {
                        "op": "update",
                        "line_id": consume_line.pk,
                        "fields": {"quantity_kg": new_actual},
                    }
                )

        if line_changes:
            edit_movement(
                movement=consume_movement,
                changes={},
                line_changes=line_changes,
                reason=f"míchání: skutečná spotřeba (dávka #{mixing_job.pk})",
                user=user,
            )

        # Produce Movement — single-line prijem from Míchárna supplier.
        produce_movement = Movement(
            branch=mixing_job.branch,
            kind=Movement.Kind.PRIJEM,
            date_issued=date_issued,
            dodavatel=_micharna_supplier(),
            note=(
                f"Míchání směsi {mixing_job.mixture.name_cs} — vyrobeno "
                f"{actual_produced_qty} kg (cíl {mixing_job.target_qty} kg)."
            ),
        )
        produce_line = MovementLine(
            product=mixing_job.mixture,
            quantity_kg=actual_produced_qty,
        )
        if actual_produced_qty > 0:
            apply_movement(
                movement=produce_movement,
                lines=[produce_line],
                user=user,
            )
            mixing_job.produce_movement = produce_movement

        mixing_job.actual_produced_qty = actual_produced_qty
        mixing_job.state = MixingJob.State.DONE
        from django.utils.timezone import now as _now
        mixing_job.finished_at = _now()
        mixing_job.save(
            update_fields=[
                "actual_produced_qty",
                "state",
                "finished_at",
                "produce_movement",
            ]
        )
        # Refresh consume_movement reference (in case edit_movement
        # changed the line set / quantities).
        _ = mixing_job.consume_movement
        return mixing_job


def cancel_mixing_job(
    *,
    mixing_job: MixingJob,
    reason: str,
    user,
) -> MixingJob:
    """Cancel a PLANNED or RUNNING job per
    [0044](../context/decisions/0044-reservations-planned-states.md):

    - PLANNED → CANCELLED: no consume_movement exists yet, nothing to
      reverse. Just mark the state and finished_at.
    - RUNNING → CANCELLED: zero each consume line via edit_movement
      (returns consumed stock to the branch).
    """
    if mixing_job.state not in {
        MixingJob.State.PLANNED,
        MixingJob.State.RUNNING,
    }:
        raise ValidationError(
            {"state": "Lze zrušit pouze plánovanou nebo probíhající dávku."}
        )
    if not reason or not reason.strip():
        raise ValidationError(
            {"reason": "Důvod zrušení je povinný."}
        )

    with transaction.atomic():
        if mixing_job.state == MixingJob.State.RUNNING:
            consume_movement = mixing_job.consume_movement
            # Remove every line of the consume Movement — edit_movement
            # reverses the stock delta atomically and writes a LINE_REMOVED
            # audit row per line.
            line_changes = [
                {"op": "remove", "line_id": ml.pk}
                for ml in consume_movement.lines.all()
            ]
            if line_changes:
                edit_movement(
                    movement=consume_movement,
                    changes={},
                    line_changes=line_changes,
                    reason=f"míchání zrušeno: {reason}",
                    user=user,
                )

        mixing_job.state = MixingJob.State.CANCELLED
        mixing_job.cancel_reason = reason
        from django.utils.timezone import now as _now
        mixing_job.finished_at = _now()
        mixing_job.save(
            update_fields=["state", "cancel_reason", "finished_at"]
        )
        return mixing_job


def record_completed_mixing_job(
    *,
    branch,
    mixture: Product,
    target_qty: Decimal,
    actual_produced_qty: Decimal,
    line_actuals_by_component_pk: dict[int, Decimal] | None = None,
    user,
    as_of=None,
    note: str = "",
    sarze_by_component: dict | None = None,
) -> MixingJob:
    """One-shot path per 0039: start + finish in a single transaction
    using a single `as_of` date for both Movements. The operator uses
    this when they forgot to open the screen at start and is recording
    a completed batch after the fact.
    """
    job = start_mixing_job(
        branch=branch,
        mixture=mixture,
        target_qty=target_qty,
        user=user,
        as_of=as_of,
        note=note,
        sarze_by_component=sarze_by_component,
    )
    line_actuals: dict[int, Decimal] = {}
    if line_actuals_by_component_pk:
        for jl in job.lines.all():
            if jl.component_product_id in line_actuals_by_component_pk:
                line_actuals[jl.pk] = line_actuals_by_component_pk[
                    jl.component_product_id
                ]
    return finish_mixing_job(
        mixing_job=job,
        actual_produced_qty=actual_produced_qty,
        line_actuals=line_actuals,
        user=user,
        as_of=as_of,
    )


# ---------------------------------------------------------------------------
# Manual stock adjustment (Pass 5d, per decision 0041)
# ---------------------------------------------------------------------------


_ADJUSTMENT_COUNTERPARTY_NAME = "Inventura / ruční úprava"
_ADJUSTMENT_NOTE_PREFIX = "[STAV] "


def _adjustment_supplier() -> Supplier:
    return Supplier.objects.get(
        name=_ADJUSTMENT_COUNTERPARTY_NAME, is_internal=True
    )


def _adjustment_customer() -> Customer:
    return Customer.objects.get(
        name=_ADJUSTMENT_COUNTERPARTY_NAME, is_internal=True
    )


def apply_stock_adjustment(
    *,
    product: Product,
    branch: Branch,
    new_quantity: Decimal,
    reason: str,
    user,
    as_of: date | None = None,
) -> Movement | None:
    """Bring `Stock(product, branch).quantity` to `new_quantity` by
    writing one synthetic Movement.

    Per [0041](../context/decisions/0041-manual-stock-adjustment.md):
    every Stock delta goes through `apply_movement`, never raw
    UPDATE. The delta determines the kind:
        delta > 0 → prijem from internal "Inventura / ruční úprava"
        delta < 0 → vydej to   internal "Inventura / ruční úprava"
        delta = 0 → noop (returns None)
    The Movement's note is `"[STAV] " + reason` so it shows up
    cleanly in Historie and can be filtered out by a future
    inventura-tab.

    The internal counterparty has `is_internal=True` so the dodák
    hook in apply_movement is skipped (no PDF, no e-mail).
    """
    from decimal import Decimal as _D
    if not reason or not reason.strip():
        raise ValidationError(
            {"reason": "Důvod ruční úpravy stavu je povinný."}
        )

    new_quantity = _D(new_quantity)
    if new_quantity < 0:
        raise ValidationError(
            {"new_quantity": "Stav nemůže být záporný."}
        )

    current = Stock.objects.filter(product=product, branch=branch).first()
    current_qty = current.quantity if current else _D("0.000")
    delta = new_quantity - current_qty
    if delta == 0:
        return None

    issue_date = as_of or date.today()
    clean_reason = reason.strip()
    note = f"{_ADJUSTMENT_NOTE_PREFIX}{clean_reason}"

    if delta > 0:
        # Stock up: synthetic prijem from internal supplier.
        movement = Movement(
            branch=branch,
            kind=Movement.Kind.PRIJEM,
            dodavatel=_adjustment_supplier(),
            date_issued=issue_date,
            note=note,
            created_by=user,
        )
    else:
        # Stock down: synthetic vydej to internal customer.
        movement = Movement(
            branch=branch,
            kind=Movement.Kind.VYDEJ,
            odberatel=_adjustment_customer(),
            date_issued=issue_date,
            note=note,
            created_by=user,
        )

    line = MovementLine(product=product, quantity_kg=abs(delta))
    return apply_movement(movement=movement, lines=[line], user=user)


# ---------------------------------------------------------------------------
# Reorder threshold + reservations + low-stock summary
# (per decisions 0043 + 0044 + 0045)
# ---------------------------------------------------------------------------


@dataclass
class LowStockRow:
    """One (product, branch) row below threshold for the dashboard panel
    and the daily summary e-mail."""

    product: Product
    branch: Branch
    on_hand: Decimal
    reserved: Decimal
    effective: Decimal
    threshold: Decimal
    deficit: Decimal


def threshold_for(product: Product, branch: Branch) -> Decimal | None:
    """Return the per-(product, branch) reorder threshold per 0043.

    Lookup order: branch-specific override row, then the product
    default, then None (no alert).
    """
    override = StockThresholdOverride.objects.filter(
        product=product, branch=branch
    ).first()
    if override is not None:
        return override.threshold_kg
    return product.reorder_threshold_kg


def reserved_kg(product: Product, branch: Branch) -> Decimal:
    """Sum outgoing reservations at one branch for one product per 0044.

    Two sources:
    - PLANNED MixingJobLine.derived_qty where the job runs at `branch`
      AND the line's component_product == `product`.
    - PLANNED PlannedTransfer.quantity_kg where source_branch == `branch`
      AND product == `product`.

    Does NOT subtract incoming planned transfers — reservations are
    outgoing-only in MVP. Returns Decimal("0.000") when nothing matches.
    """
    from django.db.models import Sum

    mixing_total = (
        MixingJobLine.objects.filter(
            mixing_job__state=MixingJob.State.PLANNED,
            mixing_job__branch=branch,
            component_product=product,
        ).aggregate(s=Sum("derived_qty"))["s"]
        or Decimal("0.000")
    )
    transfer_total = (
        PlannedTransfer.objects.filter(
            state=PlannedTransfer.State.PLANNED,
            source_branch=branch,
            product=product,
        ).aggregate(s=Sum("quantity_kg"))["s"]
        or Decimal("0.000")
    )
    return (mixing_total + transfer_total).quantize(Decimal("0.001"))


def effective_kg(product: Product, branch: Branch) -> Decimal:
    """Stock.quantity − reserved_kg(product, branch) per 0043.

    Returns Decimal("0.000") when no Stock row exists.
    """
    stock = Stock.objects.filter(product=product, branch=branch).first()
    on_hand = stock.quantity if stock else Decimal("0.000")
    return (on_hand - reserved_kg(product, branch)).quantize(Decimal("0.001"))


def low_stock_rows() -> list[LowStockRow]:
    """Every (product, branch) pair where a threshold is set and
    effective < threshold. Sorted by deficit DESC.

    Used by the owner dashboard panel, branch dashboard, product
    detail page, AND the daily summary e-mail. One source of truth.
    """
    rows: list[LowStockRow] = []
    branches = list(Branch.objects.filter(is_active=True).order_by("code"))
    products = list(
        Product.objects.filter(is_active=True).order_by("name_cs")
    )
    if not branches or not products:
        return rows

    # Pre-fetch stocks + overrides in two queries.
    stocks_by_pair = {
        (s.product_id, s.branch_id): s.quantity
        for s in Stock.objects.filter(
            product__in=products, branch__in=branches
        )
    }
    overrides_by_pair = {
        (o.product_id, o.branch_id): o.threshold_kg
        for o in StockThresholdOverride.objects.filter(
            product__in=products, branch__in=branches
        )
    }

    for product in products:
        for branch in branches:
            threshold = overrides_by_pair.get((product.pk, branch.pk))
            if threshold is None:
                threshold = product.reorder_threshold_kg
            if threshold is None:
                continue
            on_hand = stocks_by_pair.get(
                (product.pk, branch.pk), Decimal("0.000")
            )
            reserved = reserved_kg(product, branch)
            effective = (on_hand - reserved).quantize(Decimal("0.001"))
            if effective < threshold:
                rows.append(
                    LowStockRow(
                        product=product,
                        branch=branch,
                        on_hand=on_hand,
                        reserved=reserved,
                        effective=effective,
                        threshold=threshold,
                        deficit=(threshold - effective).quantize(Decimal("0.001")),
                    )
                )
    rows.sort(key=lambda r: r.deficit, reverse=True)
    return rows


# ---------------------------------------------------------------------------
# PlannedTransfer execution + cancellation (per 0044)
# ---------------------------------------------------------------------------


_TRANSFER_COUNTERPARTY_NAME = "Převod mezi pobočkami"


def _transfer_customer() -> Customer:
    return Customer.objects.get(
        name=_TRANSFER_COUNTERPARTY_NAME, is_internal=False
    )


def _transfer_supplier() -> Supplier:
    return Supplier.objects.get(
        name=_TRANSFER_COUNTERPARTY_NAME, is_internal=False
    )


def execute_planned_transfer(
    transfer: PlannedTransfer,
    *,
    executed_by,
    as_of: date | None = None,
) -> tuple[Movement, Movement]:
    """Run the výdej leg at source + the příjem leg at target atomically.

    Per [0044](../context/decisions/0044-reservations-planned-states.md):
    counterparty pair is `is_internal=False` so the existing dodák
    auto-issue + e-mail hook fires on the výdej leg — the dodák is the
    physical paper for the driver. Both Movements get a back-FK to
    `transfer` so the audit trail can reconstruct the pairing.

    Refuses if `transfer.state != PLANNED`.
    """
    if transfer.state != PlannedTransfer.State.PLANNED:
        raise ValidationError(
            {"state": "Provést lze pouze plánovaný převod."}
        )

    issue_date = as_of or transfer.scheduled_for or date.today()

    with transaction.atomic():
        # Source-leg výdej.
        vydej = Movement(
            branch=transfer.source_branch,
            kind=Movement.Kind.VYDEJ,
            date_issued=issue_date,
            odberatel=_transfer_customer(),
            note=(
                f"Převod {transfer.product.name_cs} "
                f"{transfer.quantity_kg} kg → "
                f"{transfer.target_branch.code} "
                f"(plán #{transfer.pk})."
            ),
            transfer=transfer,
        )
        vydej_line = MovementLine(
            product=transfer.product,
            quantity_kg=transfer.quantity_kg,
        )
        apply_movement(movement=vydej, lines=[vydej_line], user=executed_by)

        # Target-leg příjem.
        prijem = Movement(
            branch=transfer.target_branch,
            kind=Movement.Kind.PRIJEM,
            date_issued=issue_date,
            dodavatel=_transfer_supplier(),
            note=(
                f"Převod {transfer.product.name_cs} "
                f"{transfer.quantity_kg} kg ← "
                f"{transfer.source_branch.code} "
                f"(plán #{transfer.pk})."
            ),
            transfer=transfer,
        )
        prijem_line = MovementLine(
            product=transfer.product,
            quantity_kg=transfer.quantity_kg,
        )
        apply_movement(movement=prijem, lines=[prijem_line], user=executed_by)

        transfer.state = PlannedTransfer.State.DONE
        transfer.save(update_fields=["state"])
        return vydej, prijem


def cancel_planned_transfer(
    transfer: PlannedTransfer,
    *,
    cancelled_by,
) -> None:
    """Mark a PLANNED transfer as CANCELLED. No stock touched.

    Per [0044](../context/decisions/0044-reservations-planned-states.md):
    all authenticated users may cancel (Matej 2026-06-14 confirmation —
    symmetric with create). No tier gate beyond LoginRequiredMiddleware.
    The `cancelled_by` parameter is accepted for symmetry / future
    audit; not persisted in MVP.
    """
    if transfer.state != PlannedTransfer.State.PLANNED:
        raise ValidationError(
            {"state": "Zrušit lze pouze plánovaný převod."}
        )
    transfer.state = PlannedTransfer.State.CANCELLED
    transfer.save(update_fields=["state"])


# ---------------------------------------------------------------------------
# Daily low-stock summary e-mail (per 0045)
# ---------------------------------------------------------------------------


def _format_low_stock_list(rows: list[LowStockRow]) -> str:
    """Render the multi-line `<seznam>` placeholder body."""
    lines = []
    for r in rows:
        lines.append(
            f"- {r.product.name_cs} @ {r.branch.code}: "
            f"efektivně {r.effective} kg / práh {r.threshold} kg"
        )
    return "\n".join(lines)


def send_low_stock_summary() -> int | None:
    """Build the low-stock list, render templates, send to Petr per 0045.

    Returns the number of rows on success; None if nothing to send.
    Recipient is `Settings.recipient_petr`. If the recipient is blank,
    no e-mail is sent and None is returned (matches the
    `_assert_recipients_set` posture without raising).
    """
    rows = low_stock_rows()
    if not rows:
        return None

    s = Settings.load()
    if not s.recipient_petr:
        return None

    today = date.today().strftime("%d. %m. %Y")
    seznam = _format_low_stock_list(rows)
    subject = (
        s.template_low_stock_subject
        .replace("<datum>", today)
        .replace("<seznam>", seznam)
    )
    body = (
        s.template_low_stock_body
        .replace("<datum>", today)
        .replace("<seznam>", seznam)
    )

    from_email = (
        f"{s.email_from_name} <{s.email_from_address}>"
        if s.email_from_name and s.email_from_address
        else (s.email_from_address or None)
    )

    msg = EmailMessage(
        subject=subject,
        body=body,
        from_email=from_email,
        to=[s.recipient_petr],
    )
    msg.send(fail_silently=False)
    return len(rows)


# ---------------------------------------------------------------------------
# XLS recipe importer — per decision 0048
# ---------------------------------------------------------------------------

_RATIO_QUANT = Decimal("0.000001")  # 6 dp, matches RecipeComponent.ratio


@dataclass
class ParsedRecipeLine:
    """One ingredient row from an XLS recipe, post-Title-Case."""

    name_cs: str
    qty_kg: Decimal
    ratio: Decimal  # 6 dp, guaranteed > 0, sum of all == Decimal("1.000000")


@dataclass
class ParsedRecipe:
    """A whole recipe as extracted from the XLS, prior to the operator's review.

    `warnings` is shown on the review page; operator may still confirm.
    """

    mixture_name: str
    total_kg: Decimal
    notes: str
    lines: list[ParsedRecipeLine]
    warnings: list[str]


def _title_case_cs(text: str) -> str:
    """Czech-friendly Title Case: first letter of each whitespace word capitalised,
    the rest lower-cased. Preserves diacritics. ``str.title()`` from stdlib mangles
    apostrophes / digits but is fine for ingredient names like "KRUPIČKA"."""
    return " ".join(
        word[:1].upper() + word[1:].lower() if word else word
        for word in text.split(" ")
    )


def _normalize_ratios(quantities: list[Decimal]) -> list[Decimal]:
    """Compute ratio = qty / total quantised to 6 dp; normalise the *largest*
    so the sum is exactly ``Decimal("1.000000")``. Tiebreak: first equal one wins.

    Raises ``ValidationError`` (Czech) if any ratio quantises to 0 — that would
    violate the model's ``ratio > 0`` constraint at commit time.
    """
    total = sum(quantities, Decimal("0"))
    if total <= 0:
        raise ValidationError("Součet hmotností surovin musí být > 0.")
    ratios = [(q / total).quantize(_RATIO_QUANT) for q in quantities]
    drift = Decimal("1.000000") - sum(ratios, Decimal("0"))
    if drift != 0:
        # Absorb rounding drift into the largest ratio.
        max_idx = max(range(len(ratios)), key=lambda i: ratios[i])
        ratios[max_idx] = (ratios[max_idx] + drift).quantize(_RATIO_QUANT)
    return ratios


def _parse_xls_rows(file_obj, filename: str) -> list[list[Any]]:
    """Dispatch by extension and return rows as a uniform list[list[Any]].

    Both libraries called in **in-memory mode** — no filesystem path is ever
    passed, only bytes / file-like objects. We rewind ``file_obj`` first so
    Django's UploadedFile (after FormField validation) gives us the full payload.
    """
    name_lc = filename.lower()
    try:
        file_obj.seek(0)
    except (AttributeError, OSError):
        pass

    if name_lc.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dep is declared
            raise ValueError("openpyxl není nainstalován.") from exc
        wb = load_workbook(file_obj, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        rows = [
            [cell.value for cell in row]
            for row in ws.iter_rows()
        ]
        wb.close()
        return rows

    if name_lc.endswith(".xls"):
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - dep is declared
            raise ValueError("xlrd není nainstalován.") from exc
        wb = xlrd.open_workbook(file_contents=file_obj.read())
        ws = wb.sheet_by_index(0)
        return [
            [ws.cell_value(r, c) for c in range(ws.ncols)]
            for r in range(ws.nrows)
        ]

    raise ValueError("Soubor nelze přečíst — očekáván formát .xls nebo .xlsx.")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_number(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        d = Decimal(str(value))
    except Exception:
        return None
    return d if d > 0 else None


def parse_recipe_xls(file_obj, filename: str) -> ParsedRecipe:
    """Parse Petr's recipe XLS layout into a ParsedRecipe.

    Expected layout (see decision 0048):

    - Row 0, col 0 → mixture name (free-text Czech).
    - Row 1 → header (``druh suroviny`` / ``množství kg`` / …); ignored.
    - Rows 2..N → ingredient rows: col 0 = name (text), col 1 = qty kg
      (number > 0). Empty rows tolerated.
    - Row N → ``CELKEM`` row (col 0 == "CELKEM"); col 1 = stated total.
    - Rows N+ → free-form notes (every non-empty col-0 string joined by ``\n``).

    Tolerated quirks: missing CELKEM (warning added, total computed from sum);
    Title Case applied to mixture name + ingredient names.

    Raises ``ValueError`` (Czech) for unparseable input.
    """
    rows = _parse_xls_rows(file_obj, filename)
    if not rows:
        raise ValueError("Soubor je prázdný.")

    mixture_name_raw = _cell_text(rows[0][0]) if rows[0] else ""
    if not mixture_name_raw:
        raise ValueError("Soubor neobsahuje název směsi v prvním řádku.")
    mixture_name = _title_case_cs(mixture_name_raw)

    # Find the CELKEM row, if any.
    celkem_idx: int | None = None
    for i, row in enumerate(rows[1:], start=1):
        if row and _cell_text(row[0]).upper() == "CELKEM":
            celkem_idx = i
            break

    warnings: list[str] = []

    # Ingredient rows live between row 1 (header) and CELKEM (or EOF).
    ingredient_end = celkem_idx if celkem_idx is not None else len(rows)
    ingredient_names: list[str] = []
    ingredient_quantities: list[Decimal] = []
    for row in rows[2:ingredient_end]:
        if not row:
            continue
        name = _cell_text(row[0])
        qty = _cell_number(row[1]) if len(row) > 1 else None
        if not name or qty is None:
            continue
        ingredient_names.append(_title_case_cs(name))
        ingredient_quantities.append(qty)

    if not ingredient_names:
        raise ValueError("Receptura je prázdná — žádné suroviny.")

    # Total kg: prefer CELKEM, fall back to sum with a warning.
    total_kg: Decimal
    if celkem_idx is not None:
        stated_total = _cell_number(rows[celkem_idx][1]) if len(rows[celkem_idx]) > 1 else None
        if stated_total is None:
            total_kg = sum(ingredient_quantities, Decimal("0"))
            warnings.append(
                "Řádek CELKEM nemá množství — použit součet hmotností surovin."
            )
        else:
            total_kg = stated_total
    else:
        total_kg = sum(ingredient_quantities, Decimal("0"))
        warnings.append(
            "V souboru chybí řádek CELKEM — použit součet hmotností surovin."
        )

    # Notes = non-empty col-0 strings after CELKEM.
    notes_lines: list[str] = []
    if celkem_idx is not None:
        for row in rows[celkem_idx + 1:]:
            if not row:
                continue
            text = _cell_text(row[0])
            if text and text.upper() != "CELKEM":
                notes_lines.append(text)
    notes = "\n".join(notes_lines)

    # Compute & normalise ratios; reject zero-ratio edge cases.
    ratios = _normalize_ratios(ingredient_quantities)
    for name, ratio in zip(ingredient_names, ratios, strict=True):
        if ratio <= 0:
            raise ValidationError(
                f'Surovina „{name}“ má příliš malý poměr '
                "(< 0.000001 z celkové hmotnosti). Zvyšte množství v XLS "
                "nebo přidejte další desetinné místo."
            )

    parsed_lines = [
        ParsedRecipeLine(name_cs=n, qty_kg=q, ratio=r)
        for n, q, r in zip(
            ingredient_names, ingredient_quantities, ratios, strict=True
        )
    ]

    return ParsedRecipe(
        mixture_name=mixture_name,
        total_kg=total_kg,
        notes=notes,
        lines=parsed_lines,
        warnings=warnings,
    )


def create_mixture_from_review(
    *,
    header_data: dict,
    line_data: list[dict],
    user,
) -> Product:
    """Commit the operator-reviewed import as one atomic transaction.

    ``header_data`` carries ``name_cs``, ``notes``, ``total_kg`` (from the
    review header form). ``line_data`` is a list of ``{name_cs, qty_kg,
    existing_product_id}`` dicts (DELETE rows already filtered out by the
    view).

    Behaviour:

    - Refuse if an active Product with the same name already exists (iexact).
    - Pre-fetch active RAW_SPICE products once; dedupe via ``.casefold()``
      so case-variants resolve to the same row.
    - Auto-create RAW_SPICE Products for unmatched ingredient names.
    - Re-compute ratios from the (possibly edited) quantities and normalise
      so the sum is ``Decimal("1.000000")`` — same logic as parse.
    - Call ``RecipeComponent.full_clean()`` per row to enforce model
      invariants (no self-reference, mixture parent must be MIXTURE).

    Returns the newly created mixture Product.
    """
    name_cs = header_data["name_cs"].strip()
    if not name_cs:
        raise ValidationError({"name_cs": "Název směsi je povinný."})
    if not line_data:
        raise ValidationError("Receptura musí mít alespoň jednu surovinu.")

    with transaction.atomic():
        if Product.objects.filter(name_cs__iexact=name_cs, is_active=True).exists():
            raise ValidationError(
                f'Směs s názvem „{name_cs}“ už v katalogu existuje.'
            )

        # Pre-fetch all active raw spices in one query, indexed by casefolded name.
        existing_by_key = {
            p.name_cs.casefold(): p
            for p in Product.objects.filter(
                kind=Product.Kind.RAW_SPICE, is_active=True,
            )
        }

        quantities = [Decimal(str(line["qty_kg"])) for line in line_data]
        ratios = _normalize_ratios(quantities)
        for line, ratio in zip(line_data, ratios, strict=True):
            if ratio <= 0:
                raise ValidationError(
                    f'Surovina „{line["name_cs"]}“ má příliš malý poměr '
                    "(< 0.000001). Zvyšte množství nebo odeberte tuto řádku."
                )

        mixture = Product.objects.create(
            name_cs=name_cs,
            kind=Product.Kind.MIXTURE,
            notes=header_data.get("notes", "") or "",
            is_active=True,
        )

        seen_components: set[int] = set()
        for line, ratio in zip(line_data, ratios, strict=True):
            line_name = line["name_cs"].strip()
            if not line_name:
                raise ValidationError("Surovina bez názvu — vyplňte název nebo odeberte řádku.")

            existing_id = line.get("existing_product_id")
            if existing_id:
                try:
                    component_product = Product.objects.get(
                        pk=existing_id, kind=Product.Kind.RAW_SPICE
                    )
                except Product.DoesNotExist as exc:
                    raise ValidationError(
                        f'Surovina „{line_name}“ nebyla nalezena v katalogu.'
                    ) from exc
            else:
                cached = existing_by_key.get(line_name.casefold())
                if cached is not None:
                    component_product = cached
                else:
                    component_product, _ = Product.objects.get_or_create(
                        name_cs=line_name,
                        defaults={
                            "kind": Product.Kind.RAW_SPICE,
                            "is_active": True,
                        },
                    )
                    existing_by_key[line_name.casefold()] = component_product

            if component_product.pk in seen_components:
                raise ValidationError(
                    f'Surovina „{component_product.name_cs}“ je v receptuře dvakrát.'
                )
            seen_components.add(component_product.pk)

            component = RecipeComponent(
                mixture_product=mixture,
                component_product=component_product,
                ratio=ratio,
            )
            component.full_clean()
            component.save()

        return mixture
