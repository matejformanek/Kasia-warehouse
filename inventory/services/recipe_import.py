"""XLS recipe import + ratio helpers."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from typing import Any

from django.core.exceptions import ValidationError
from django.db import transaction

from ..models import (
    Product,
    RecipeComponent,
)

_RATIO_QUANT = Decimal("0.000001")  # 6 dp, matches RecipeComponent.ratio


@dataclass
class ParsedRecipeLine:
    """One ingredient row from an XLS recipe, post-Title-Case."""

    name_cs: str
    qty_kg: Decimal
    ratio: Decimal  # 6 dp, guaranteed > 0, sum of all == Decimal("1.000000")


@dataclass
class ParsedRecipe:
    """A whole recipe as extracted from the XLS, prior to the operator's review.

    `warnings` is shown on the review page; operator may still confirm.
    """

    mixture_name: str
    total_kg: Decimal
    notes: str
    lines: list[ParsedRecipeLine]
    warnings: list[str]


def _title_case_cs(text: str) -> str:
    """Czech-friendly Title Case: first letter of each whitespace word capitalised,
    the rest lower-cased. Preserves diacritics. ``str.title()`` from stdlib mangles
    apostrophes / digits but is fine for ingredient names like "KRUPIČKA"."""
    return " ".join(
        word[:1].upper() + word[1:].lower() if word else word
        for word in text.split(" ")
    )


def _normalize_ratios(quantities: list[Decimal]) -> list[Decimal]:
    """Compute ratio = qty / total quantised to 6 dp; normalise the *largest*
    so the sum is exactly ``Decimal("1.000000")``. Tiebreak: first equal one wins.

    Raises ``ValidationError`` (Czech) if any ratio quantises to 0 — that would
    violate the model's ``ratio > 0`` constraint at commit time.
    """
    total = sum(quantities, Decimal("0"))
    if total <= 0:
        raise ValidationError("Součet hmotností surovin musí být > 0.")
    ratios = [(q / total).quantize(_RATIO_QUANT) for q in quantities]
    drift = Decimal("1.000000") - sum(ratios, Decimal("0"))
    if drift != 0:
        # Absorb rounding drift into the largest ratio.
        max_idx = max(range(len(ratios)), key=lambda i: ratios[i])
        ratios[max_idx] = (ratios[max_idx] + drift).quantize(_RATIO_QUANT)
    return ratios


def _parse_xls_rows(file_obj, filename: str) -> list[list[Any]]:
    """Dispatch by extension and return rows as a uniform list[list[Any]].

    Both libraries called in **in-memory mode** — no filesystem path is ever
    passed, only bytes / file-like objects. We rewind ``file_obj`` first so
    Django's UploadedFile (after FormField validation) gives us the full payload.
    """
    name_lc = filename.lower()
    try:
        file_obj.seek(0)
    except (AttributeError, OSError):
        pass

    if name_lc.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dep is declared
            raise ValueError("openpyxl není nainstalován.") from exc
        wb = load_workbook(file_obj, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        rows = [
            [cell.value for cell in row]
            for row in ws.iter_rows()
        ]
        wb.close()
        return rows

    if name_lc.endswith(".xls"):
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - dep is declared
            raise ValueError("xlrd není nainstalován.") from exc
        wb = xlrd.open_workbook(file_contents=file_obj.read())
        ws = wb.sheet_by_index(0)
        return [
            [ws.cell_value(r, c) for c in range(ws.ncols)]
            for r in range(ws.nrows)
        ]

    raise ValueError("Soubor nelze přečíst — očekáván formát .xls nebo .xlsx.")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_number(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        d = Decimal(str(value))
    except Exception:
        return None
    return d if d > 0 else None


def parse_recipe_xls(file_obj, filename: str) -> ParsedRecipe:
    """Parse Petr's recipe XLS layout into a ParsedRecipe.

    Expected layout (see decision 0048):

    - Row 0, col 0 → mixture name (free-text Czech).
    - Row 1 → header (``druh suroviny`` / ``množství kg`` / …); ignored.
    - Rows 2..N → ingredient rows: col 0 = name (text), col 1 = qty kg
      (number > 0). Empty rows tolerated.
    - Row N → ``CELKEM`` row (col 0 == "CELKEM"); col 1 = stated total.
    - Rows N+ → free-form notes (every non-empty col-0 string joined by ``\n``).

    Tolerated quirks: missing CELKEM (warning added, total computed from sum);
    Title Case applied to mixture name + ingredient names.

    Raises ``ValueError`` (Czech) for unparseable input.
    """
    rows = _parse_xls_rows(file_obj, filename)
    if not rows:
        raise ValueError("Soubor je prázdný.")

    mixture_name_raw = _cell_text(rows[0][0]) if rows[0] else ""
    if not mixture_name_raw:
        raise ValueError("Soubor neobsahuje název směsi v prvním řádku.")
    mixture_name = _title_case_cs(mixture_name_raw)

    # Find the CELKEM row, if any.
    celkem_idx: int | None = None
    for i, row in enumerate(rows[1:], start=1):
        if row and _cell_text(row[0]).upper() == "CELKEM":
            celkem_idx = i
            break

    warnings: list[str] = []

    # Ingredient rows live between row 1 (header) and CELKEM (or EOF).
    ingredient_end = celkem_idx if celkem_idx is not None else len(rows)
    ingredient_names: list[str] = []
    ingredient_quantities: list[Decimal] = []
    for row in rows[2:ingredient_end]:
        if not row:
            continue
        name = _cell_text(row[0])
        qty = _cell_number(row[1]) if len(row) > 1 else None
        if not name or qty is None:
            continue
        ingredient_names.append(_title_case_cs(name))
        ingredient_quantities.append(qty)

    if not ingredient_names:
        raise ValueError("Receptura je prázdná — žádné suroviny.")

    # Total kg: prefer CELKEM, fall back to sum with a warning.
    total_kg: Decimal
    if celkem_idx is not None:
        stated_total = _cell_number(rows[celkem_idx][1]) if len(rows[celkem_idx]) > 1 else None
        if stated_total is None:
            total_kg = sum(ingredient_quantities, Decimal("0"))
            warnings.append(
                "Řádek CELKEM nemá množství — použit součet hmotností surovin."
            )
        else:
            total_kg = stated_total
    else:
        total_kg = sum(ingredient_quantities, Decimal("0"))
        warnings.append(
            "V souboru chybí řádek CELKEM — použit součet hmotností surovin."
        )

    # Notes = non-empty col-0 strings after CELKEM.
    notes_lines: list[str] = []
    if celkem_idx is not None:
        for row in rows[celkem_idx + 1:]:
            if not row:
                continue
            text = _cell_text(row[0])
            if text and text.upper() != "CELKEM":
                notes_lines.append(text)
    notes = "\n".join(notes_lines)

    # Compute & normalise ratios; reject zero-ratio edge cases.
    ratios = _normalize_ratios(ingredient_quantities)
    for name, ratio in zip(ingredient_names, ratios, strict=True):
        if ratio <= 0:
            raise ValidationError(
                f'Surovina „{name}“ má příliš malý poměr '
                "(< 0.000001 z celkové hmotnosti). Zvyšte množství v XLS "
                "nebo přidejte další desetinné místo."
            )

    parsed_lines = [
        ParsedRecipeLine(name_cs=n, qty_kg=q, ratio=r)
        for n, q, r in zip(
            ingredient_names, ingredient_quantities, ratios, strict=True
        )
    ]

    return ParsedRecipe(
        mixture_name=mixture_name,
        total_kg=total_kg,
        notes=notes,
        lines=parsed_lines,
        warnings=warnings,
    )


def create_mixture_from_review(
    *,
    header_data: dict,
    line_data: list[dict],
    user,
) -> Product:
    """Commit the operator-reviewed import as one atomic transaction.

    ``header_data`` carries ``name_cs``, ``notes``, ``total_kg`` (from the
    review header form). ``line_data`` is a list of ``{name_cs, qty_kg,
    existing_product_id}`` dicts (DELETE rows already filtered out by the
    view).

    Behaviour:

    - Refuse if an active Product with the same name already exists (iexact).
    - Pre-fetch active RAW_SPICE products once; dedupe via ``.casefold()``
      so case-variants resolve to the same row.
    - Auto-create RAW_SPICE Products for unmatched ingredient names.
    - Re-compute ratios from the (possibly edited) quantities and normalise
      so the sum is ``Decimal("1.000000")`` — same logic as parse.
    - Call ``RecipeComponent.full_clean()`` per row to enforce model
      invariants (no self-reference, mixture parent must be MIXTURE).

    Returns the newly created mixture Product.
    """
    name_cs = header_data["name_cs"].strip()
    if not name_cs:
        raise ValidationError({"name_cs": "Název směsi je povinný."})
    if not line_data:
        raise ValidationError("Receptura musí mít alespoň jednu surovinu.")

    with transaction.atomic():
        if Product.objects.filter(name_cs__iexact=name_cs, is_active=True).exists():
            raise ValidationError(
                f'Směs s názvem „{name_cs}“ už v katalogu existuje.'
            )

        # Pre-fetch all active raw spices in one query, indexed by casefolded name.
        existing_by_key = {
            p.name_cs.casefold(): p
            for p in Product.objects.filter(
                kind=Product.Kind.RAW_SPICE, is_active=True,
            )
        }

        quantities = [Decimal(str(line["qty_kg"])) for line in line_data]
        ratios = _normalize_ratios(quantities)
        for line, ratio in zip(line_data, ratios, strict=True):
            if ratio <= 0:
                raise ValidationError(
                    f'Surovina „{line["name_cs"]}“ má příliš malý poměr '
                    "(< 0.000001). Zvyšte množství nebo odeberte tuto řádku."
                )

        mixture = Product.objects.create(
            name_cs=name_cs,
            kind=Product.Kind.MIXTURE,
            notes=header_data.get("notes", "") or "",
            is_active=True,
        )

        seen_components: set[int] = set()
        for line, ratio in zip(line_data, ratios, strict=True):
            line_name = line["name_cs"].strip()
            if not line_name:
                raise ValidationError("Surovina bez názvu — vyplňte název nebo odeberte řádku.")

            existing_id = line.get("existing_product_id")
            if existing_id:
                try:
                    component_product = Product.objects.get(
                        pk=existing_id, kind=Product.Kind.RAW_SPICE
                    )
                except Product.DoesNotExist as exc:
                    raise ValidationError(
                        f'Surovina „{line_name}“ nebyla nalezena v katalogu.'
                    ) from exc
            else:
                cached = existing_by_key.get(line_name.casefold())
                if cached is not None:
                    component_product = cached
                else:
                    component_product, _ = Product.objects.get_or_create(
                        name_cs=line_name,
                        defaults={
                            "kind": Product.Kind.RAW_SPICE,
                            "is_active": True,
                        },
                    )
                    existing_by_key[line_name.casefold()] = component_product

            if component_product.pk in seen_components:
                raise ValidationError(
                    f'Surovina „{component_product.name_cs}“ je v receptuře dvakrát.'
                )
            seen_components.add(component_product.pk)

            component = RecipeComponent(
                mixture_product=mixture,
                component_product=component_product,
                ratio=ratio,
            )
            component.full_clean()
            component.save()

        return mixture
